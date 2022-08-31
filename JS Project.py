from tkinter import *
import tkinter as tk
from tkinter import ttk
from tkinter.ttk import Combobox
from PIL import ImageTk, Image
import openpyxl
from openpyxl import Workbook
import os
import csv
import pandas as pd
from scrape_linkedin import ProfileScraper
import json
from math import*

mydata = []
main=Tk()
t1 = StringVar()
t2 = StringVar()

# Geometory
main.title("JS Project")
main.geometry("1100x500")
main.config(highlightbackground="black",highlightthickness= 6)

def update(rows):
    global mydata
    mydata = rows
    trv.delete(*trv.get_children())
    for i in rows:
        if i[0] == "2":
            trv.insert('','end',value=i[1], tags = "unchecked")
              
        elif i[0] == "1":
            trv.insert('','end',value=i[1], text = "Section---------------------->")
            i[1] = style.configure("mystyle.Treeview" , font=('Calibri', 13,'bold'))

        else:
            trv.insert('','end', text = "")

def toggleCheck(event):
    rowid = trv.identify_row(event.y)
    tag = trv.item(rowid,"tags")[0]
    tags = list(trv.item(rowid,"tags"))
    tags.remove(tag)
    trv.item(rowid,tags=tags)
    if tag == "checked":
        trv.item(rowid,tags="unchecked")
    else:
        trv.item(rowid,tags="checked")
        if tag == "unchecked":
            item = trv.item(trv.focus())
            with open('CheckedKeyword.csv', 'a', newline='') as csvfile:
                writer = csv.writer(csvfile, delimiter=' ')
                writer.writerow(item['values'])

# Function for  Checked and Unchecked Images
def checked():
    im_checked = ImageTk.PhotoImage(Image.open("D:/PROJECT/1) MAIN/BACK/checked.png"))
    im_unchecked = ImageTk.PhotoImage(Image.open("D:/PROJECT/1) MAIN/BACK/unchecked.png"))

class Test:
    def __init__(self, tk):
        self.var = StringVar()
        self.lable = ttk.Label(wrapper1,text="Roll").place(x=20,y=50)
        self.data = [os.path.splitext(filename)[0] for filename in os.listdir("D:/PROJECT/1) MAIN/BACK/Users/")]
        
        # Combobox
        self.cb = Combobox(main, values = self.data)
        self.cb.place(x=150, y=75)
        self.cb.bind('<<ComboboxSelected>>', self.select)
        
    def select(self, tk):
        get_value = self.cb.get() 
        with open("D:/PROJECT/1) MAIN/BACK/Users/" + get_value + ".csv") as myfile:
            csvread = csv.reader(myfile, delimiter=",")
            for i in csvread:
                mydata.append(i)
        update(mydata)
        read_file = pd.read_csv (r'D:/PROJECT/1) MAIN/BACK/Users/' + get_value + '.csv')
        read_file.to_excel (r'D:/PROJECT/1) MAIN/BACK/' + get_value + '.xlsx', index = None, header=True)

        # Submit Button
        self.sbtn = Button(wrapper2, text = "Submit", command = self.submit)
        self.sbtn.place(x=270,y=250)
        self.sbtn.bind('<<Button 1>>', self.submit)
        
    # Submit Function
    def submit(event):
        roll = event.cb.get()
        company_name = t1.get()

        # Creating Row_count of every section
        book = openpyxl.load_workbook(roll + '.xlsx')
        p = openpyxl.Workbook()
        sheet = book.active
        sheet1 = p.active
        row_count = sheet.max_row
        count_sum =0
        for i in range(2, row_count+1):
            row_cell = sheet.cell(row= i, column=1).value
            row_cell1 = sheet.cell(row= i, column=2).value
            
            if row_cell ==2:
                count_sum = count_sum + 1
            elif row_cell ==1:
                count_row = sheet1.cell(row=i, column=2)
                count_row.value = count_sum
                count_row1 = sheet1.cell(row=i, column=1)
                count_row1.value = row_cell1
                count_sum = 0

            rows_to_delete = [None, '', ' ',0]
            for i in p.sheetnames:
                sheet1 = p[i]
                column_b = range(1, sheet1.max_row)
                for i in reversed(column_b):
                    if sheet1.cell(i, 1).value in rows_to_delete:
                        sheet1.cell(i,1).row
                        sheet1.delete_rows(sheet1.cell(i,1).row)
                    if sheet1.cell(i, 2).value in rows_to_delete:
                        sheet1.cell(i,2).row
                        sheet1.delete_rows(sheet1.cell(i,2).row)
        p.save(roll + 'row_value.xlsx')
         
        # Creating Master Excel file
        book = openpyxl.load_workbook(roll + '.xlsx')
        p = openpyxl.Workbook()
        sheet = book.active
        sheet1 = p.active
        sheet = book.worksheets[0]
        row_count = sheet.max_row
        book_count = openpyxl.load_workbook(roll + 'row_value.xlsx')
        sheet2 = book_count.active

        c1 = sheet1.cell(row= 1, column=1)
        c1.value = 'Level'
        c2 = sheet1.cell(row= 1, column=2)
        c2.value = 'Keyword'
        c3 = sheet1.cell(row= 1, column=3)
        c3.value = 'Present'
        vr = 1
        flag = 0
        for roll_rows in range(1, row_count+1):
            cell = sheet.cell(row= roll_rows, column=1).value
            cell1 = sheet.cell(row= roll_rows, column=2).value
            cell2 = sheet.cell(row= roll_rows, column=3).value
            if cell == 1 and cell1 != "Generic":
                flag = 1
            if flag == 0:

                if cell == 1:
                    c1 = sheet1.cell(row= roll_rows, column=1)
                    c1.value = cell                    
                    c2 = sheet1.cell(row= roll_rows, column=2)
                    c2.value = cell1                    
                    c3 = sheet1.cell(row= roll_rows, column=3)
                    c3.value = cell2

                elif cell == 2:
                    with open('CheckedKeyword.csv','r') as f:
                        f1 = f.read()

                        if str(cell1) in f1: 
                            c1 = sheet1.cell(row= roll_rows, column=1)
                            c1.value = cell                    
                            c2 = sheet1.cell(row= roll_rows, column=2)
                            c2.value = cell1
                            c3 = sheet1.cell(row= roll_rows, column=3)
                            c3.value = 1

                        else:
                            c1 = sheet1.cell(row= roll_rows, column=1)
                            c1.value = cell                    
                            c2 = sheet1.cell(row= roll_rows, column=2)
                            c2.value = cell1                    
                            c3 = sheet1.cell(row= roll_rows, column=3)
                            c3.value = 0
            else:             
                if cell == 1:
                    c1 = sheet1.cell(row= roll_rows, column=1)
                    c1.value = cell                    
                    c2 = sheet1.cell(row= roll_rows, column=2)
                    c2.value = cell1                   
                    c3 = sheet1.cell(row= roll_rows, column=3)
                    c3.value = cell2
                    vr+=1

                elif cell == 2:
                    with open('CheckedKeyword.csv','r') as f:
                        f1 = f.read()
                        
                        if str(cell1) in f1: 
                            c1 = sheet1.cell(row= roll_rows, column=1)
                            c1.value = cell                    
                            c2 = sheet1.cell(row= roll_rows, column=2)
                            c2.value = cell1
                            row_cell1 = sheet2.cell(row= vr, column=2).value
                            c3 = sheet1.cell(row= roll_rows, column=3)
                            c3.value = row_cell1

                        else:
                            c1 = sheet1.cell(row= roll_rows, column=1)
                            c1.value = cell                    
                            c2 = sheet1.cell(row= roll_rows, column=2)
                            c2.value = cell1                    
                            c3 = sheet1.cell(row= roll_rows, column=3)
                            c3.value = 1
        p.save(company_name + " " + roll + ".xlsx")

        # Creating file of sum score of Master Excel file
        book = openpyxl.load_workbook(company_name + " " + roll + ".xlsx")
        p = openpyxl.Workbook()
        sheet = book.active
        sheet1 = p.active
        row_count = sheet.max_row
        row_sum = 0
        for i in range(2, row_count+1):
            row_cell = sheet.cell(row= i, column=1).value
            row_cell2 = sheet.cell(row=i, column=2).value
            row_cell1 = sheet.cell(row= i, column=3).value

            if row_cell == 1:
                count_row = sheet1.cell(row=i, column=1)
                count_row.value = row_cell2
                count_row = sheet1.cell(row=i, column=2)
                count_row.value = row_sum
                row_sum = 0
            if row_cell ==2:
                row_sum = row_sum + row_cell1

            rows_to_delete = [None, '', ' ']
            for i in p.sheetnames:
                sheet1 = p[i]
                column_b = range(1, sheet1.max_row)
                for i in reversed(column_b):
                    if sheet1.cell(i, 1).value in rows_to_delete:
                        sheet1.cell(i,1).row
                        sheet1.delete_rows(sheet1.cell(i,1).row)
        p.save(roll + 'Row_sum.xlsx')

        # Searching LinkedIn Profiles Link
        # try:
        #     from googlesearch import search
        # except ImportError:
        #     print("No module named 'google' found")
        # query = "site:in.linkedin.com you know in common,"+ "SRE AND ( Python OR Java)"+")"

        # for j in search(query, tld="co.in", num=10, pause=2):
        #     print(j)
        #     with open(roll +"Profile Links.txt","a") as f:
        #         f.write(j+"\n")
        
        # Searching LinkedIn Profiles Link
        with open("links of profiles.txt","r") as data:
            for i in range(1,295):
                name = data.readline().replace("\n","")
                if __name__ == '__main__':
                    n = 27  
                    p_name = name[n:]
                    
                # with ProfileScraper(cookie='AQEDATfA4GYERHirAAABfgFSKd8AAAF-fR0ktE0AR1aNjvGjKm4iMbZ35t30dnewo8bYBvtVOupTqmLbBIDDd5jObDhJkJ2lqXbvoBSQgJOPpXinChjo4cqXUwvkuA-azbH1sQD_DBK1A4noytits6U4') as scraper:
                #     profile = scraper.scrape(user= p_name)
                    
                # with open('D:/PROJECT/1) MAIN/BACK/a/' + p_name + '.json', 'a') as profile_file:
                #     profile_file.write(json.dumps(profile.to_dict()))
        
                # Creating Customer Score File
                book = openpyxl.load_workbook(company_name + " " + roll + ".xlsx")
                p = openpyxl.Workbook()
                sheet = book.active
                sheet1 = p.active
                sheet = book.worksheets[0]
                row_count = sheet.max_row

                c1 = sheet1.cell(row= 1, column=1)
                c1.value = 'Level'

                c2 = sheet1.cell(row= 1, column=2)
                c2.value = 'Keyword'

                c3 = sheet1.cell(row= 1, column=3)
                c3.value = 'Present'

                for roll_rows in range(1, row_count+1):
                    cell = sheet.cell(row= roll_rows, column=1).value
                    cell1 = sheet.cell(row= roll_rows, column=2).value
                    cell2 = sheet.cell(row= roll_rows, column=3).value
                    
                    if cell == 1:

                        c1 = sheet1.cell(row= roll_rows, column=1)
                        c1.value = cell
                        
                        c2 = sheet1.cell(row= roll_rows, column=2)
                        c2.value = cell1
                        
                        c3 = sheet1.cell(row= roll_rows, column=3)
                        c3.value = cell2

                    elif cell == 2:
                        with open('D:/PROJECT/1) MAIN/BACK/New_Scraped Data/' + p_name + '.json','r') as f:
                            f1 = f.read()
                            
                            if str(cell1) in f1: 

                                c1 = sheet1.cell(row= roll_rows, column=1)
                                c1.value = cell
                        
                                c2 = sheet1.cell(row= roll_rows, column=2)
                                c2.value = cell1

                                c3 = sheet1.cell(row= roll_rows, column=3)
                                c3.value = cell2

                            else:
                                c1 = sheet1.cell(row= roll_rows, column=1)
                                c1.value = cell
                        
                                c2 = sheet1.cell(row= roll_rows, column=2)
                                c2.value = cell1
                        
                                c3 = sheet1.cell(row= roll_rows, column=3)
                                c3.value = 0
                p.save('D:/PROJECT/1) MAIN/BACK/b/'  + p_name + '.xlsx')

                book = openpyxl.load_workbook('D:/PROJECT/1) MAIN/BACK/b/'  + p_name + '.xlsx')
                p = openpyxl.Workbook()
                sheet = book.active
                sheet1 = p.active
                row_count = sheet.max_row
                row_sum = 0

                for i in range(2, row_count+1):
                    row_cell = sheet.cell(row= i, column=1).value
                    row_cell2 = sheet.cell(row=i, column=2).value
                    row_cell1 = sheet.cell(row= i, column=3).value

                    if row_cell == 1:
                        count_row = sheet1.cell(row=i, column=1)
                        count_row.value = row_cell2
                        count_row = sheet1.cell(row=i, column=2)
                        count_row.value = row_sum
                        row_sum = 0
                    if row_cell ==2:
                        row_sum = row_sum + row_cell1

                    rows_to_delete = [None, '', ' ']
                    for i in p.sheetnames:
                        sheet1 = p[i]
                        column_b = range(1, sheet1.max_row)
                        for i in reversed(column_b):
                            if sheet1.cell(i, 1).value in rows_to_delete:
                                sheet1.cell(i,1).row
                                sheet1.delete_rows(sheet1.cell(i,1).row)

                p.save('D:/PROJECT/1) MAIN/BACK/c/' + p_name + '.xlsx')

                book1 = openpyxl.load_workbook('D:/PROJECT/1) MAIN/BACK/c/' + p_name + '.xlsx')
                book = openpyxl.load_workbook('SRERow_sum.xlsx')
                sheet1 = book1.active
                row_count1 = sheet1.max_row
                sheet = book.active
                row_count = sheet.max_row
                temp_list = []
                temp_list1 = [p_name]
                temp_list3 = []
                for i in range(2, row_count+1):
                    row_cell = sheet.cell(row= i, column=2).value
                    row_cell1 = sheet1.cell(row= i, column=2).value
                    row_cell2 = sheet1.cell(row= i, column=1).value
                    final_score = row_cell1 / row_cell
                    temp_list.append(final_score)
                    
                with open('D:/PROJECT/1) MAIN/BACK/d/scores.csv','a', newline='') as csvfile:
                    writer = csv.writer(csvfile)
                    writer.writerow(temp_list) 
                
                with open('D:/PROJECT/1) MAIN/BACK/d/Names.csv','a', newline='') as csvfile1:
                    writer = csv.writer(csvfile1)
                    for val in temp_list1:
                        writer.writerow([val]) 

                with open('D:/PROJECT/1) MAIN/BACK/d/Names.csv', 'r') as p:
                    #reads csv into a list of lists
                    my_list = [list(map(str,rec)) for rec in csv.reader(p, delimiter=',')]
                    
                    for row in my_list:
                        temp_list3.append(row)
                        with open('D:/PROJECT/1) MAIN/BACK/d/Final_Score_File.csv','w',newline = '') as csvfile:
                            writer = csv.writer(csvfile)
                            writer.writerow(["Name"])
                            for val in temp_list3:
                                writer.writerow(val)

        # Finding Distance
        temp_list = []
        with open('D:/PROJECT/1) MAIN/BACK/d/scores.csv', 'r') as p:
                #reads csv into a list of lists
            my_list = [list(map(float,rec)) for rec in csv.reader(p, delimiter=',')]
            
            for row in my_list:
                y = row
                # print(y)
                x = [1,1,1,1,1,1,1,1,1,1,1,1,1,1,1]
                def euclidean_distance(x,y):
                    # print(y)
                    return sqrt(sum(pow(a-b,2) for a, b in zip(x, y)))

                final = euclidean_distance(x,y)
                temp_list.append(final)
                # print(final)            
                    
                with open('D:/PROJECT/1) MAIN/BACK/d/Distance.csv','w',newline = '') as csvfile:
                    writer = csv.writer(csvfile)
                    writer.writerow(["Name", "Scores"])
                    for val in temp_list:
                        writer.writerow([val])

        read_file = pd.read_csv (r'D:/PROJECT/1) MAIN/BACK/d/Distance.csv')
        read_file.to_excel (r'D:/PROJECT/1) MAIN/BACK/d/Distance.xlsx', index = None, header=True)
        read_file = pd.read_csv (r'D:/PROJECT/1) MAIN/BACK/d/Final_Score_File.csv')
        read_file.to_excel (r'D:/PROJECT/1) MAIN/BACK/d/Final_Score_File.xlsx', index = None, header=True)

        book = openpyxl.load_workbook('D:/PROJECT/1) MAIN/BACK/d/Distance.xlsx')
        sheet = book.active
        sheet = book.worksheets[0]
        row_count = sheet.max_row
        book1 = openpyxl.load_workbook('D:/PROJECT/1) MAIN/BACK/d/Final_Score_File.xlsx')
        sheet2 = book1.active
        p = openpyxl.Workbook()
        sheet1 = p.active

        c1 = sheet1.cell(row= 1, column=1)
        c1.value = 'Name'
        c2 = sheet1.cell(row= 1, column=2)
        c2.value = 'Scores'

        for roll_rows in range(2, row_count+1):    
            cell = sheet2.cell(row= roll_rows, column=1).value
            cell2 = sheet.cell(row= roll_rows, column=1).value

            c3 = sheet1.cell(row= roll_rows, column=1)
            c3.value = cell
            # print(cell)
            c4 = sheet1.cell(row= roll_rows, column=2)
            c4.value = cell2
            
        p.save('D:/PROJECT/1) MAIN/BACK/Final/Before_Sorting.xlsx')

        # Sorting of Distance calculated file
        read_file = pd.read_excel (r'D:/PROJECT/1) MAIN/BACK/Final/Before_Sorting.xlsx')
        read_file.to_csv (r'D:/PROJECT/1) MAIN/BACK/Final/Before_Sorting.csv', index = None, header=True)
        cdv = pd.read_csv("D:/PROJECT/1) MAIN/BACK/Final/Before_Sorting.csv")
        Sorted_Scores = cdv.sort_values(["Scores"])
        Sorted_Scores.to_csv('D:/PROJECT/1) MAIN/BACK/Final/Final List.csv', index=False)
            
#Company Details Section
wrapper1=LabelFrame(main,text="Details")
wrapper1.pack(fill="both",expand="yes",padx=20,pady=10)
Label(wrapper1,text="Company Name").place(x=20,y=10)
name = Entry(wrapper1, textvariable = t1)
name.place(x=130,y=12)

# Inputs Section
wrapper2=LabelFrame(main,text="Skills")
wrapper2.pack(fill="both",expand="yes",padx=20,pady=10)

im_checked = ImageTk.PhotoImage(Image.open("D:/PROJECT/1) MAIN/BACK/checked.png"))
im_unchecked = ImageTk.PhotoImage(Image.open("D:/PROJECT/1) MAIN/BACK/unchecked.png"))

trv = ttk.Treeview(wrapper2,columns=(1,2,3,4),style="mystyle.Treeview")
style = ttk.Style(trv)
style.configure("mystyle.Treeview.Heading", font=('Calibri', 15,'bold'))
trv.tag_configure('checked', image = im_checked)
trv.tag_configure('unchecked', image = im_unchecked)
trv.pack()

trv.heading('#0', text ="")
trv.heading('#1', text="Keyword")
trv.bind('<Button 1>', toggleCheck)

get_class = Test(main)
main.mainloop()